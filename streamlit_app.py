import streamlit as st
import sqlite3
import pandas as pd
import json
import re
import os
from datetime import datetime, date, timedelta
from io import BytesIO
import hashlib
import secrets

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CA Client Master",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = "ca_practice.db"
AUTH_YAML = "users.json"

INDIAN_STATES = [
    "Andaman and Nicobar Islands", "Andhra Pradesh", "Arunachal Pradesh",
    "Assam", "Bihar", "Chandigarh", "Chhattisgarh", "Dadra and Nagar Haveli and Daman and Diu",
    "Delhi", "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jammu and Kashmir",
    "Jharkhand", "Karnataka", "Kerala", "Ladakh", "Lakshadweep", "Madhya Pradesh",
    "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha",
    "Puducherry", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana",
    "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
]

CONSTITUTIONS = [
    "Individual", "HUF", "Proprietorship", "Partnership", "LLP",
    "Private Limited", "Public Limited", "OPC", "Trust", "AOP", "BOI", "Others",
]

COMPANY_CONSTITUTIONS = ["LLP", "Private Limited", "Public Limited", "OPC"]

CLIENT_STATUSES = ["Active", "Inactive", "Prospect", "Discontinued"]
RISK_FLAGS = ["Low", "Medium", "High"]
TAX_REGIMES = ["Old", "New"]
RESIDENTIAL_STATUSES = ["Resident", "NRI", "RNOR"]
ITR_FORMS = ["ITR-1", "ITR-2", "ITR-3", "ITR-4", "ITR-5", "ITR-6", "ITR-7"]
GST_REG_TYPES = ["Regular", "Composition", "SEZ", "Casual", "Non-resident", "ISD", "TDS-TCS", "Not Registered"]
GSTR_FREQUENCY = ["Monthly", "Quarterly (QRMP)"]
GSTR_RETURNS = ["GSTR-1", "GSTR-3B", "GSTR-9", "GSTR-9C"]
TDS_FORMS = ["24Q", "26Q", "27Q", "27EQ"]
ROC_FORMS = ["AOC-4", "MGT-7", "ADT-1", "DIR-3 KYC"]
PAYMENT_TERMS = ["Advance", "Monthly", "Quarterly", "On Completion"]
INVOICE_FREQ = ["Monthly", "Per Service", "Annual"]
ACCOUNT_TYPES = ["Savings", "Current", "Overdraft"]
BOOKING_FREQ = ["Monthly", "Quarterly", "Annual"]
AUDIT_FREQ = ["Monthly", "Quarterly", "Annual"]
CLIENT_SOURCES = ["Referral", "Walk-in", "Social Media", "ICAI", "Other"]
CLIENT_IMPORTANCE = ["Key Account", "Regular", "Small", "One-time"]
PAYMENT_BEHAVIOUR = ["Prompt", "Delayed (30d)", "Requires Follow-up", "Defaulter"]
CONTACT_DESIG = ["Accountant", "CFO", "Director", "Owner", "Partner", "Other"]

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Lato:wght@300;400;700&display=swap');

    html, body, [class*="css"] { font-family: 'Lato', sans-serif; }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f1923 0%, #1a2a3a 100%);
        border-right: 1px solid #2d4a6b;
    }
    [data-testid="stSidebar"] * { color: #d4e6f1 !important; }
    [data-testid="stSidebar"] .stRadio label { 
        padding: 6px 0; font-size: 0.95rem; letter-spacing: 0.02em;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2 {
        color: #e8b86d !important; font-family: 'Playfair Display', serif !important;
        font-size: 1.1rem; letter-spacing: 0.05em;
    }

    /* Main area */
    .main .block-container { padding-top: 1.5rem; }
    h1 { font-family: 'Playfair Display', serif !important; color: #1a2a3a !important; }
    h2, h3 { font-family: 'Playfair Display', serif !important; color: #1a2a3a !important; }

    /* Metric cards */
    [data-testid="metric-container"] {
        background: #fff;
        border: 1px solid #e8ecf0;
        border-left: 4px solid #e8b86d;
        border-radius: 6px;
        padding: 12px 16px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    [data-testid="metric-container"] label { color: #5a7a96 !important; font-size: 0.78rem; text-transform: uppercase; letter-spacing: 0.08em; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] { color: #1a2a3a !important; font-weight: 700; }

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] { gap: 4px; background: #f0f4f8; border-radius: 8px; padding: 4px; }
    .stTabs [data-baseweb="tab"] { border-radius: 6px; font-size: 0.82rem; padding: 6px 12px; }
    .stTabs [aria-selected="true"] { background: #1a2a3a !important; color: #e8b86d !important; }

    /* Buttons */
    .stButton > button {
        background: #1a2a3a; color: #e8b86d; border: 1px solid #e8b86d;
        border-radius: 5px; font-weight: 600; letter-spacing: 0.04em;
        transition: all 0.2s;
    }
    .stButton > button:hover { background: #e8b86d; color: #1a2a3a; }

    /* Badge helpers */
    .badge-low   { background:#d4edda; color:#155724; padding:2px 10px; border-radius:12px; font-size:0.78rem; font-weight:600; }
    .badge-med   { background:#fff3cd; color:#856404; padding:2px 10px; border-radius:12px; font-size:0.78rem; font-weight:600; }
    .badge-high  { background:#f8d7da; color:#721c24; padding:2px 10px; border-radius:12px; font-size:0.78rem; font-weight:600; }
    .badge-active{ background:#cce5ff; color:#004085; padding:2px 10px; border-radius:12px; font-size:0.78rem; font-weight:600; }

    /* Section header */
    .section-hdr {
        background: linear-gradient(90deg, #1a2a3a, #2d4a6b);
        color: #e8b86d !important; padding: 8px 16px; border-radius: 5px;
        font-family: 'Playfair Display', serif; font-size: 0.95rem;
        margin: 12px 0 8px 0; letter-spacing: 0.04em;
    }

    /* Page title */
    .page-title {
        font-family: 'Playfair Display', serif;
        font-size: 1.7rem; font-weight: 700;
        color: #1a2a3a; border-bottom: 2px solid #e8b86d;
        padding-bottom: 8px; margin-bottom: 20px;
    }

    /* Dataframe */
    [data-testid="stDataFrame"] { border: 1px solid #e8ecf0; border-radius: 6px; }

    /* Validation inline */
    .val-ok  { color: #28a745; font-size: 0.8rem; }
    .val-err { color: #dc3545; font-size: 0.8rem; }

    /* Alert table */
    .alert-row { background: #fff8e1; }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def get_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_connection()
    conn.execute("""
    CREATE TABLE IF NOT EXISTS clients (
        -- A: Identity
        client_id TEXT PRIMARY KEY,
        client_full_name TEXT NOT NULL,
        constitution TEXT,
        date_of_birth_incorp TEXT,
        father_promoter_name TEXT,
        reg_residential_addr TEXT,
        city TEXT,
        state TEXT,
        pincode TEXT,
        client_status TEXT DEFAULT 'Active',
        client_since TEXT,
        source_of_client TEXT,
        -- B: Contact
        primary_mobile TEXT,
        alternate_mobile TEXT,
        primary_email TEXT,
        alternate_email TEXT,
        whatsapp_number TEXT,
        contact_person_name TEXT,
        contact_person_desig TEXT,
        contact_person_mobile TEXT,
        contact_person_email TEXT,
        -- C: Income Tax
        pan TEXT,
        aadhaar_number TEXT,
        residential_status TEXT,
        tax_regime TEXT,
        it_portal_username TEXT,
        it_portal_password TEXT DEFAULT 'Stored in password manager',
        traces_registered INTEGER DEFAULT 0,
        tan TEXT,
        tan_applicable INTEGER DEFAULT 0,
        itr_form_type TEXT,
        tax_audit_44ab INTEGER DEFAULT 0,
        advance_tax_tracking INTEGER DEFAULT 0,
        tds_return_filing INTEGER DEFAULT 0,
        tds_forms_applicable TEXT,
        -- D: GST
        gstin TEXT,
        gst_registration_type TEXT,
        gst_registration_date TEXT,
        gst_annual_turnover REAL DEFAULT 0,
        gstr_filing_frequency TEXT,
        gst_portal_username TEXT,
        gst_portal_password TEXT DEFAULT 'Stored in password manager',
        gst_practitioner_auth INTEGER DEFAULT 0,
        gst_returns_in_scope TEXT,
        gstr_9c_recon INTEGER DEFAULT 0,
        e_way_bill_filing INTEGER DEFAULT 0,
        -- E: ROC/MCA
        cin_llpin TEXT,
        date_of_incorporation TEXT,
        registered_office_addr TEXT,
        authorised_capital REAL DEFAULT 0,
        paidup_capital REAL DEFAULT 0,
        number_of_directors INTEGER DEFAULT 0,
        director_names_din TEXT,
        mca_portal_username TEXT,
        mca_portal_password TEXT DEFAULT 'Stored in password manager',
        financial_year_end TEXT DEFAULT 'March 31',
        agm_date_current_year TEXT,
        auditor_appt_date TEXT,
        auditor_tenure_yrs INTEGER DEFAULT 0,
        previous_auditor_name TEXT,
        nature_of_business TEXT,
        caro_applicable INTEGER DEFAULT 0,
        listed_entity INTEGER DEFAULT 0,
        roc_forms_in_scope TEXT,
        -- F: Services
        itr_filing INTEGER DEFAULT 0,
        gst_compliance INTEGER DEFAULT 0,
        statutory_audit INTEGER DEFAULT 0,
        tax_audit_3cd INTEGER DEFAULT 0,
        roc_mca_compliance INTEGER DEFAULT 0,
        bookkeeping INTEGER DEFAULT 0,
        bookkeeping_frequency TEXT,
        payroll_processing INTEGER DEFAULT 0,
        internal_audit INTEGER DEFAULT 0,
        internal_audit_freq TEXT,
        tds_return_service INTEGER DEFAULT 0,
        fema_rbi_advisory INTEGER DEFAULT 0,
        valuation_services INTEGER DEFAULT 0,
        -- G: Engagement
        el_issued INTEGER DEFAULT 0,
        el_issue_date TEXT,
        el_signed_by_client INTEGER DEFAULT 0,
        el_last_renewed TEXT,
        scope_last_updated TEXT,
        kyc_docs_collected INTEGER DEFAULT 0,
        onboarding_complete INTEGER DEFAULT 0,
        poa_auth_letter INTEGER DEFAULT 0,
        -- H: Compliance Tracker
        last_itr_ay TEXT,
        itr_ack_number TEXT,
        last_gst_return TEXT,
        last_tds_return TEXT,
        last_roc_filing TEXT,
        last_roc_form_filed TEXT,
        pending_with_client TEXT,
        pending_with_ca TEXT,
        it_notices_scrutiny TEXT,
        gst_notices TEXT,
        other_proceedings TEXT,
        -- I: Fees
        itr_filing_fee REAL DEFAULT 0,
        gst_monthly_retainer REAL DEFAULT 0,
        gstr9_9c_fee REAL DEFAULT 0,
        tds_return_fee REAL DEFAULT 0,
        statutory_audit_fee REAL DEFAULT 0,
        tax_audit_3cd_fee REAL DEFAULT 0,
        roc_mca_annual_fee REAL DEFAULT 0,
        bookkeeping_monthly REAL DEFAULT 0,
        payroll_fee REAL DEFAULT 0,
        other_services_fee REAL DEFAULT 0,
        total_annual_fee REAL DEFAULT 0,
        gst_on_fees INTEGER DEFAULT 0,
        payment_terms TEXT,
        invoice_frequency TEXT,
        last_invoice_date TEXT,
        last_invoice_amount REAL DEFAULT 0,
        total_billed_fy REAL DEFAULT 0,
        total_received_fy REAL DEFAULT 0,
        outstanding_balance REAL DEFAULT 0,
        fee_revision_due TEXT,
        fee_notes TEXT,
        -- J: Bank
        bank_name TEXT,
        account_number TEXT,
        ifsc_code TEXT,
        account_type TEXT,
        account_holder_name TEXT,
        bank_linked_to_pan INTEGER DEFAULT 0,
        no_of_additional_accts INTEGER DEFAULT 0,
        bank2_name TEXT,
        bank2_acc_no TEXT,
        bank2_ifsc TEXT,
        -- K: Other Registrations
        pf_account_number TEXT,
        pf_establishment_code TEXT,
        esi_number TEXT,
        msme_udyam_reg_no TEXT,
        import_export_code TEXT,
        -- L: Risk & CRM
        risk_flag TEXT DEFAULT 'Low',
        payment_behaviour TEXT,
        referred_by TEXT,
        clients_referred TEXT,
        client_importance TEXT,
        next_review_date TEXT,
        last_contacted TEXT,
        next_followup_date TEXT,
        followup_purpose TEXT,
        -- M: Internal
        drive_folder_link TEXT,
        compuoffice_code TEXT,
        tallyprime_name TEXT,
        internal_notes TEXT,
        exit_discontinue_reason TEXT,
        -- Meta
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    conn.commit()


def generate_client_id():
    conn = get_connection()
    year = datetime.now().year
    cur = conn.execute(
        "SELECT client_id FROM clients WHERE client_id LIKE ? ORDER BY client_id DESC LIMIT 1",
        (f"CA-{year}-%",)
    )
    row = cur.fetchone()
    if row:
        last_num = int(row["client_id"].split("-")[-1])
        return f"CA-{year}-{last_num + 1:04d}"
    return f"CA-{year}-0001"


def insert_client(data: dict) -> str:
    conn = get_connection()
    data["client_id"] = generate_client_id()
    data["created_at"] = datetime.now().isoformat()
    data["updated_at"] = datetime.now().isoformat()
    cols = ", ".join(data.keys())
    placeholders = ", ".join(["?" for _ in data])
    try:
        conn.execute(f"INSERT INTO clients ({cols}) VALUES ({placeholders})", list(data.values()))
        conn.commit()
        return data["client_id"]
    except Exception as e:
        raise e


def update_client(data: dict):
    conn = get_connection()
    data["updated_at"] = datetime.now().isoformat()
    client_id = data.pop("client_id")
    set_clause = ", ".join([f"{k} = ?" for k in data.keys()])
    vals = list(data.values()) + [client_id]
    conn.execute(f"UPDATE clients SET {set_clause} WHERE client_id = ?", vals)
    conn.commit()
    data["client_id"] = client_id


def delete_client(client_id: str):
    conn = get_connection()
    conn.execute("DELETE FROM clients WHERE client_id = ?", (client_id,))
    conn.commit()


@st.cache_data(ttl=5)
def get_all_clients() -> pd.DataFrame:
    conn = get_connection()
    return pd.read_sql("SELECT * FROM clients ORDER BY client_full_name", conn)


@st.cache_data(ttl=5)
def get_summary_view() -> pd.DataFrame:
    conn = get_connection()
    cols = """client_id, client_full_name, constitution, pan, gstin,
              primary_mobile, primary_email, itr_filing, gst_compliance,
              statutory_audit, roc_mca_compliance, bookkeeping,
              total_annual_fee, el_signed_by_client, last_itr_ay,
              last_gst_return, last_roc_filing, outstanding_balance,
              client_status, risk_flag, next_review_date, drive_folder_link"""
    return pd.read_sql(f"SELECT {cols} FROM clients ORDER BY client_full_name", conn)


def get_client_by_id(client_id: str) -> dict:
    conn = get_connection()
    cur = conn.execute("SELECT * FROM clients WHERE client_id = ?", (client_id,))
    row = cur.fetchone()
    return dict(row) if row else {}


def search_clients(filters: dict) -> pd.DataFrame:
    conn = get_connection()
    where_parts = ["1=1"]
    params = []
    if filters.get("text"):
        t = f"%{filters['text']}%"
        where_parts.append("(client_full_name LIKE ? OR pan LIKE ? OR gstin LIKE ? OR primary_mobile LIKE ? OR primary_email LIKE ?)")
        params.extend([t, t, t, t, t])
    for col in ["client_status", "risk_flag", "client_importance"]:
        vals = filters.get(col, [])
        if vals:
            placeholders = ",".join(["?" for _ in vals])
            where_parts.append(f"{col} IN ({placeholders})")
            params.extend(vals)
    for col in ["constitution"]:
        vals = filters.get(col, [])
        if vals:
            placeholders = ",".join(["?" for _ in vals])
            where_parts.append(f"{col} IN ({placeholders})")
            params.extend(vals)
    for svc in ["itr_filing", "gst_compliance", "statutory_audit", "roc_mca_compliance", "bookkeeping"]:
        if filters.get(svc):
            where_parts.append(f"{svc} = 1")
    if filters.get("outstanding_only"):
        where_parts.append("outstanding_balance > 0")
    if filters.get("el_signed") == "Yes":
        where_parts.append("el_signed_by_client = 1")
    elif filters.get("el_signed") == "No":
        where_parts.append("el_signed_by_client = 0")
    if filters.get("state"):
        vals = filters["state"]
        if vals:
            placeholders = ",".join(["?" for _ in vals])
            where_parts.append(f"state IN ({placeholders})")
            params.extend(vals)
    if filters.get("city"):
        where_parts.append("LOWER(city) LIKE ?")
        params.append(f"%{filters['city'].lower()}%")
    if filters.get("fee_min") is not None:
        where_parts.append("total_annual_fee >= ?")
        params.append(filters["fee_min"])
    if filters.get("fee_max") is not None:
        where_parts.append("total_annual_fee <= ?")
        params.append(filters["fee_max"])
    sql = f"SELECT * FROM clients WHERE {' AND '.join(where_parts)} ORDER BY client_full_name"
    return pd.read_sql(sql, conn, params=params)


def calc_total_fee(d: dict) -> float:
    fee_cols = ["itr_filing_fee","gst_monthly_retainer","gstr9_9c_fee","tds_return_fee",
                "statutory_audit_fee","tax_audit_3cd_fee","roc_mca_annual_fee",
                "bookkeeping_monthly","payroll_fee","other_services_fee"]
    return sum(float(d.get(c, 0) or 0) for c in fee_cols)


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def validate_pan(pan: str) -> bool:
    return bool(re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan.strip().upper())) if pan else True

def validate_gstin(g: str) -> bool:
    return bool(re.match(r'^\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[0-9A-Z]$', g.strip().upper())) if g else True

def validate_mobile(m: str) -> bool:
    return bool(re.match(r'^[6-9]\d{9}$', m.strip())) if m else True

def validate_pincode(p: str) -> bool:
    return bool(re.match(r'^\d{6}$', p.strip())) if p else True

def validate_ifsc(i: str) -> bool:
    return bool(re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', i.strip().upper())) if i else True

def mask_aadhaar(a: str) -> str:
    digits = re.sub(r'\D', '', a)
    if len(digits) == 12:
        return f"XXXX-XXXX-{digits[-4:]}"
    return a


# ─────────────────────────────────────────────────────────────────────────────
# SAMPLE DATA
# ─────────────────────────────────────────────────────────────────────────────
def seed_sample_data():
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    if count > 0:
        return

    samples = [
        {
            "client_full_name": "Rajesh Kumar Sharma",
            "constitution": "Individual",
            "date_of_birth_incorp": "1980-03-15",
            "father_promoter_name": "Suresh Kumar Sharma",
            "reg_residential_addr": "12, Navrangpura Society",
            "city": "Ahmedabad",
            "state": "Gujarat",
            "pincode": "380009",
            "client_status": "Active",
            "client_since": "2019-04-01",
            "source_of_client": "Referral",
            "primary_mobile": "9876543210",
            "primary_email": "rajesh.sharma@gmail.com",
            "pan": "ABCPS1234R",
            "residential_status": "Resident",
            "tax_regime": "New",
            "it_portal_username": "ABCPS1234R",
            "it_portal_password": "Stored in password manager",
            "itr_form_type": "ITR-1",
            "itr_filing": 1,
            "gst_compliance": 0,
            "statutory_audit": 0,
            "roc_mca_compliance": 0,
            "bookkeeping": 0,
            "el_issued": 1,
            "el_signed_by_client": 1,
            "el_issue_date": "2024-04-01",
            "kyc_docs_collected": 1,
            "onboarding_complete": 1,
            "last_itr_ay": "AY 2024-25",
            "itr_filing_fee": 3000.0,
            "total_annual_fee": 3000.0,
            "total_billed_fy": 3000.0,
            "total_received_fy": 3000.0,
            "outstanding_balance": 0.0,
            "bank_name": "HDFC Bank",
            "account_type": "Savings",
            "bank_linked_to_pan": 1,
            "risk_flag": "Low",
            "payment_behaviour": "Prompt",
            "client_importance": "Regular",
            "next_review_date": "2025-04-01",
            "referred_by": "Mahesh Patel",
            "drive_folder_link": "https://drive.google.com/drive/folders/sample1",
        },
        {
            "client_full_name": "Priya Textiles (Prop: Priya Shah)",
            "constitution": "Proprietorship",
            "date_of_birth_incorp": "1975-07-22",
            "father_promoter_name": "Hasmukh Shah",
            "reg_residential_addr": "Shop 7, Textile Market, Raipur",
            "city": "Surat",
            "state": "Gujarat",
            "pincode": "395003",
            "client_status": "Active",
            "client_since": "2020-06-15",
            "source_of_client": "Walk-in",
            "primary_mobile": "9712345678",
            "primary_email": "priya.textiles@gmail.com",
            "pan": "BCZPS5678K",
            "gstin": "24BCZPS5678K1ZA",
            "gst_registration_type": "Regular",
            "gst_registration_date": "2017-07-01",
            "gst_annual_turnover": 4500000.0,
            "gstr_filing_frequency": "Monthly",
            "gst_portal_username": "24BCZPS5678K1ZA",
            "gst_portal_password": "Stored in password manager",
            "gst_returns_in_scope": json.dumps(["GSTR-1", "GSTR-3B", "GSTR-9"]),
            "itr_form_type": "ITR-3",
            "itr_filing": 1,
            "gst_compliance": 1,
            "statutory_audit": 0,
            "roc_mca_compliance": 0,
            "bookkeeping": 1,
            "bookkeeping_frequency": "Monthly",
            "el_issued": 1,
            "el_signed_by_client": 1,
            "el_issue_date": "2024-04-01",
            "kyc_docs_collected": 1,
            "onboarding_complete": 1,
            "last_itr_ay": "AY 2024-25",
            "last_gst_return": "GSTR-3B Mar 2025",
            "itr_filing_fee": 5000.0,
            "gst_monthly_retainer": 2500.0,
            "bookkeeping_monthly": 3000.0,
            "total_annual_fee": 90000.0,
            "total_billed_fy": 90000.0,
            "total_received_fy": 75000.0,
            "outstanding_balance": 15000.0,
            "risk_flag": "Low",
            "payment_behaviour": "Prompt",
            "client_importance": "Regular",
            "next_review_date": "2025-04-01",
            "drive_folder_link": "https://drive.google.com/drive/folders/sample2",
        },
        {
            "client_full_name": "M/s Mehta & Associates",
            "constitution": "Partnership",
            "father_promoter_name": "Kiran Mehta (Managing Partner)",
            "reg_residential_addr": "201, Commerce House, CG Road",
            "city": "Ahmedabad",
            "state": "Gujarat",
            "pincode": "380006",
            "client_status": "Active",
            "client_since": "2018-01-01",
            "source_of_client": "ICAI",
            "primary_mobile": "9824567890",
            "primary_email": "mehta.associates@outlook.com",
            "pan": "AAMFM4321B",
            "gstin": "24AAMFM4321B1ZC",
            "gst_registration_type": "Regular",
            "gst_registration_date": "2017-08-01",
            "gst_annual_turnover": 12000000.0,
            "gstr_filing_frequency": "Monthly",
            "gst_portal_username": "24AAMFM4321B1ZC",
            "gst_portal_password": "Stored in password manager",
            "gst_returns_in_scope": json.dumps(["GSTR-1", "GSTR-3B", "GSTR-9"]),
            "itr_form_type": "ITR-5",
            "itr_filing": 1,
            "gst_compliance": 1,
            "statutory_audit": 0,
            "roc_mca_compliance": 0,
            "bookkeeping": 1,
            "bookkeeping_frequency": "Monthly",
            "contact_person_name": "Deepak Mehta",
            "contact_person_desig": "Partner",
            "contact_person_mobile": "9824567891",
            "el_issued": 1,
            "el_signed_by_client": 1,
            "el_issue_date": "2024-04-01",
            "kyc_docs_collected": 1,
            "onboarding_complete": 1,
            "last_itr_ay": "AY 2024-25",
            "last_gst_return": "GSTR-3B Mar 2025",
            "itr_filing_fee": 8000.0,
            "gst_monthly_retainer": 5000.0,
            "bookkeeping_monthly": 8000.0,
            "total_annual_fee": 156000.0,
            "total_billed_fy": 156000.0,
            "total_received_fy": 156000.0,
            "outstanding_balance": 0.0,
            "risk_flag": "Low",
            "payment_behaviour": "Prompt",
            "client_importance": "Key Account",
            "next_review_date": "2025-04-01",
            "drive_folder_link": "https://drive.google.com/drive/folders/sample3",
        },
        {
            "client_full_name": "Ganesh Pharma Pvt Ltd",
            "constitution": "Private Limited",
            "date_of_birth_incorp": "2010-09-12",
            "father_promoter_name": "Dr. Ganesh Patel",
            "reg_residential_addr": "Plot 42, GIDC Phase II, Vatva",
            "city": "Ahmedabad",
            "state": "Gujarat",
            "pincode": "382445",
            "client_status": "Active",
            "client_since": "2015-04-01",
            "source_of_client": "Referral",
            "primary_mobile": "9998887776",
            "primary_email": "accounts@ganeshpharma.in",
            "pan": "AABCG7890P",
            "gstin": "24AABCG7890P1ZB",
            "gst_registration_type": "Regular",
            "gst_registration_date": "2017-07-01",
            "gst_annual_turnover": 85000000.0,
            "gstr_filing_frequency": "Monthly",
            "gst_portal_username": "24AABCG7890P1ZB",
            "gst_portal_password": "Stored in password manager",
            "gst_returns_in_scope": json.dumps(["GSTR-1", "GSTR-3B", "GSTR-9", "GSTR-9C"]),
            "gstr_9c_recon": 1,
            "cin_llpin": "U24230GJ2010PTC060123",
            "date_of_incorporation": "2010-09-12",
            "registered_office_addr": "Plot 42, GIDC Phase II, Vatva, Ahmedabad",
            "authorised_capital": 10000000.0,
            "paidup_capital": 5000000.0,
            "number_of_directors": 3,
            "director_names_din": json.dumps([
                {"name": "Dr. Ganesh Patel", "din": "01234567", "email": "ganesh@ganeshpharma.in", "kyc_status": "Active"},
                {"name": "Hema Patel", "din": "01234568", "email": "hema@ganeshpharma.in", "kyc_status": "Active"},
                {"name": "Rahul Patel", "din": "01234569", "email": "rahul@ganeshpharma.in", "kyc_status": "Active"},
            ]),
            "mca_portal_username": "ganeshpharma",
            "mca_portal_password": "Stored in password manager",
            "nature_of_business": "Pharmaceutical Manufacturing - NIC 21001",
            "caro_applicable": 1,
            "listed_entity": 0,
            "roc_forms_in_scope": json.dumps(["AOC-4", "MGT-7", "ADT-1", "DIR-3 KYC"]),
            "itr_form_type": "ITR-6",
            "tax_audit_44ab": 1,
            "itr_filing": 1,
            "gst_compliance": 1,
            "statutory_audit": 1,
            "tax_audit_3cd": 1,
            "roc_mca_compliance": 1,
            "bookkeeping": 1,
            "bookkeeping_frequency": "Monthly",
            "payroll_processing": 1,
            "contact_person_name": "Vinod Agrawal",
            "contact_person_desig": "CFO",
            "contact_person_mobile": "9998887775",
            "contact_person_email": "cfo@ganeshpharma.in",
            "el_issued": 1,
            "el_signed_by_client": 1,
            "el_issue_date": "2024-04-01",
            "kyc_docs_collected": 1,
            "onboarding_complete": 1,
            "last_itr_ay": "AY 2024-25",
            "last_gst_return": "GSTR-3B Mar 2025",
            "last_roc_filing": "2024-10-15",
            "last_roc_form_filed": "AOC-4",
            "itr_filing_fee": 25000.0,
            "gst_monthly_retainer": 15000.0,
            "gstr9_9c_fee": 35000.0,
            "statutory_audit_fee": 150000.0,
            "tax_audit_3cd_fee": 50000.0,
            "roc_mca_annual_fee": 30000.0,
            "bookkeeping_monthly": 20000.0,
            "payroll_fee": 8000.0,
            "total_annual_fee": 576000.0,
            "total_billed_fy": 576000.0,
            "total_received_fy": 500000.0,
            "outstanding_balance": 76000.0,
            "risk_flag": "Low",
            "payment_behaviour": "Prompt",
            "client_importance": "Key Account",
            "next_review_date": "2025-04-01",
            "drive_folder_link": "https://drive.google.com/drive/folders/sample4",
        },
        {
            "client_full_name": "TechVenture Solutions LLP",
            "constitution": "LLP",
            "date_of_birth_incorp": "2018-04-01",
            "father_promoter_name": "Nishant Joshi (Designated Partner)",
            "reg_residential_addr": "803, Iscon Emporio, SG Highway",
            "city": "Ahmedabad",
            "state": "Gujarat",
            "pincode": "380054",
            "client_status": "Active",
            "client_since": "2021-01-15",
            "source_of_client": "Social Media",
            "primary_mobile": "9510203040",
            "primary_email": "finance@techventuresllp.com",
            "pan": "AADFT1122L",
            "gstin": "24AADFT1122L1ZG",
            "gst_registration_type": "Regular",
            "gst_registration_date": "2018-05-01",
            "gst_annual_turnover": 22000000.0,
            "gstr_filing_frequency": "Monthly",
            "gst_portal_username": "24AADFT1122L1ZG",
            "gst_portal_password": "Stored in password manager",
            "gst_returns_in_scope": json.dumps(["GSTR-1", "GSTR-3B", "GSTR-9"]),
            "cin_llpin": "AAB-1234",
            "date_of_incorporation": "2018-04-01",
            "registered_office_addr": "803, Iscon Emporio, SG Highway, Ahmedabad",
            "authorised_capital": 1000000.0,
            "paidup_capital": 500000.0,
            "number_of_directors": 2,
            "director_names_din": json.dumps([
                {"name": "Nishant Joshi", "din": "08123456", "email": "nishant@techventuresllp.com", "kyc_status": "Active"},
                {"name": "Pooja Mehta", "din": "08123457", "email": "pooja@techventuresllp.com", "kyc_status": "Active"},
            ]),
            "mca_portal_username": "techventurellp",
            "mca_portal_password": "Stored in password manager",
            "nature_of_business": "IT Services & Consulting - NIC 62011",
            "caro_applicable": 0,
            "roc_forms_in_scope": json.dumps(["AOC-4", "MGT-7"]),
            "itr_form_type": "ITR-5",
            "itr_filing": 1,
            "gst_compliance": 1,
            "statutory_audit": 1,
            "roc_mca_compliance": 1,
            "bookkeeping": 1,
            "bookkeeping_frequency": "Monthly",
            "el_issued": 1,
            "el_signed_by_client": 0,
            "kyc_docs_collected": 1,
            "onboarding_complete": 1,
            "last_itr_ay": "AY 2024-25",
            "last_gst_return": "GSTR-3B Feb 2025",
            "last_roc_filing": "2024-11-20",
            "last_roc_form_filed": "MGT-7",
            "itr_filing_fee": 15000.0,
            "gst_monthly_retainer": 8000.0,
            "statutory_audit_fee": 75000.0,
            "roc_mca_annual_fee": 20000.0,
            "bookkeeping_monthly": 12000.0,
            "total_annual_fee": 254000.0,
            "total_billed_fy": 254000.0,
            "total_received_fy": 200000.0,
            "outstanding_balance": 54000.0,
            "risk_flag": "Medium",
            "payment_behaviour": "Delayed (30d)",
            "client_importance": "Key Account",
            "next_review_date": "2025-04-01",
            "drive_folder_link": "https://drive.google.com/drive/folders/sample5",
        },
    ]

    for s in samples:
        s["client_id"] = generate_client_id()
        s.setdefault("created_at", datetime.now().isoformat())
        s.setdefault("updated_at", datetime.now().isoformat())
        cols = ", ".join(s.keys())
        placeholders = ", ".join(["?" for _ in s])
        conn.execute(f"INSERT INTO clients ({cols}) VALUES ({placeholders})", list(s.values()))
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# SHARED FORM RENDERER
# ─────────────────────────────────────────────────────────────────────────────
def section_hdr(title):
    st.markdown(f'<div class="section-hdr">{title}</div>', unsafe_allow_html=True)


def render_form(defaults: dict = None, mode: str = "add") -> dict:
    """Render all 13 tabs and return collected form data dict."""
    d = defaults or {}

    def g(key, fallback=None):
        v = d.get(key, fallback)
        return v if v is not None else fallback

    def gf(key): return float(g(key, 0) or 0)
    def gi(key): return bool(int(g(key, 0) or 0))
    def gs(key, fallback=""): return str(g(key, fallback) or "")
    def gd(key):
        v = gs(key)
        if v:
            try: return datetime.strptime(v[:10], "%Y-%m-%d").date()
            except: pass
        return None

    def parse_json_list(key):
        v = gs(key)
        if not v: return []
        try: return json.loads(v)
        except: return []

    constitution = gs("constitution", "Individual")

    tab_labels = [
        "🪪 Identity", "📞 Contact", "💼 Income Tax", "🧾 GST",
        "🏢 ROC/MCA", "⚙️ Services", "📄 Engagement",
        "📅 Compliance", "💰 Fees", "🏦 Bank",
        "🔖 Registrations", "⚠️ Risk & CRM", "🗂️ Internal"
    ]
    tabs = st.tabs(tab_labels)
    form_data = {}

    # ── TAB A: Identity ──────────────────────────────────────────────────────
    with tabs[0]:
        section_hdr("🪪 Client Identity")
        if mode == "edit":
            st.text_input("Client ID", value=gs("client_id"), disabled=True)
            st.text_input("Created At", value=gs("created_at"), disabled=True)

        c1, c2 = st.columns(2)
        form_data["client_full_name"] = c1.text_input("Full Legal Name *", value=gs("client_full_name"))
        form_data["constitution"] = c2.selectbox("Constitution *", CONSTITUTIONS,
            index=CONSTITUTIONS.index(gs("constitution", "Individual")) if gs("constitution", "Individual") in CONSTITUTIONS else 0)

        c1, c2 = st.columns(2)
        dob = gd("date_of_birth_incorp")
        dob_val = c1.date_input("Date of Birth / Incorporation", value=dob, format="DD/MM/YYYY")
        form_data["date_of_birth_incorp"] = dob_val.isoformat() if dob_val else ""
        form_data["father_promoter_name"] = c2.text_input("Father's / Promoter Name", value=gs("father_promoter_name"))

        form_data["reg_residential_addr"] = st.text_area("Registered / Residential Address", value=gs("reg_residential_addr"), height=80)

        c1, c2, c3 = st.columns(3)
        form_data["city"] = c1.text_input("City", value=gs("city"))
        state_idx = INDIAN_STATES.index(gs("state", "Gujarat")) if gs("state", "Gujarat") in INDIAN_STATES else INDIAN_STATES.index("Gujarat")
        form_data["state"] = c2.selectbox("State", INDIAN_STATES, index=state_idx)
        pin = c3.text_input("Pincode", value=gs("pincode"))
        form_data["pincode"] = pin
        if pin and not validate_pincode(pin):
            c3.markdown('<span class="val-err">❌ Must be 6 digits</span>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        status_idx = CLIENT_STATUSES.index(gs("client_status", "Active")) if gs("client_status", "Active") in CLIENT_STATUSES else 0
        form_data["client_status"] = c1.selectbox("Client Status *", CLIENT_STATUSES, index=status_idx)
        since = gd("client_since")
        since_val = c2.date_input("Client Since", value=since, format="DD/MM/YYYY")
        form_data["client_since"] = since_val.isoformat() if since_val else ""

        src_idx = CLIENT_SOURCES.index(gs("source_of_client", "Referral")) if gs("source_of_client", "Referral") in CLIENT_SOURCES else 0
        form_data["source_of_client"] = st.selectbox("Source of Client", CLIENT_SOURCES, index=src_idx)

    # ── TAB B: Contact ───────────────────────────────────────────────────────
    with tabs[1]:
        section_hdr("📞 Contact Details")
        c1, c2 = st.columns(2)
        mob = c1.text_input("Primary Mobile *", value=gs("primary_mobile"))
        form_data["primary_mobile"] = mob
        if mob and not validate_mobile(mob):
            c1.markdown('<span class="val-err">❌ Must be 10 digits starting 6-9</span>', unsafe_allow_html=True)
        elif mob:
            c1.markdown('<span class="val-ok">✅ Valid</span>', unsafe_allow_html=True)
        form_data["alternate_mobile"] = c2.text_input("Alternate Mobile", value=gs("alternate_mobile"))

        c1, c2 = st.columns(2)
        form_data["primary_email"] = c1.text_input("Primary Email *", value=gs("primary_email"))
        form_data["alternate_email"] = c2.text_input("Alternate Email", value=gs("alternate_email"))
        form_data["whatsapp_number"] = st.text_input("WhatsApp Number", value=gs("whatsapp_number"))

        section_hdr("👤 Contact Person (for entities)")
        c1, c2 = st.columns(2)
        form_data["contact_person_name"] = c1.text_input("Contact Person Name", value=gs("contact_person_name"))
        desig_idx = CONTACT_DESIG.index(gs("contact_person_desig", "Accountant")) if gs("contact_person_desig", "Accountant") in CONTACT_DESIG else 0
        form_data["contact_person_desig"] = c2.selectbox("Designation", CONTACT_DESIG, index=desig_idx)
        c1, c2 = st.columns(2)
        form_data["contact_person_mobile"] = c1.text_input("Contact Person Mobile", value=gs("contact_person_mobile"))
        form_data["contact_person_email"] = c2.text_input("Contact Person Email", value=gs("contact_person_email"))

    # ── TAB C: Income Tax ────────────────────────────────────────────────────
    with tabs[2]:
        section_hdr("💼 Income Tax Details")
        c1, c2 = st.columns(2)
        pan = c1.text_input("PAN", value=gs("pan"), max_chars=10)
        form_data["pan"] = pan.upper()
        if pan:
            if validate_pan(pan):
                c1.markdown('<span class="val-ok">✅ Valid PAN format</span>', unsafe_allow_html=True)
            else:
                c1.markdown('<span class="val-err">❌ Format: AAAAA9999A</span>', unsafe_allow_html=True)
        form_data["aadhaar_number"] = c2.text_input("Aadhaar (last 4 digits shown)", value=gs("aadhaar_number"),
            help="Will be stored as XXXX-XXXX-XXXX")

        c1, c2 = st.columns(2)
        rs_idx = RESIDENTIAL_STATUSES.index(gs("residential_status", "Resident")) if gs("residential_status", "Resident") in RESIDENTIAL_STATUSES else 0
        form_data["residential_status"] = c1.selectbox("Residential Status", RESIDENTIAL_STATUSES, index=rs_idx)
        regime_idx = TAX_REGIMES.index(gs("tax_regime", "New")) if gs("tax_regime", "New") in TAX_REGIMES else 0
        form_data["tax_regime"] = c2.selectbox("Tax Regime", TAX_REGIMES, index=regime_idx)

        c1, c2 = st.columns(2)
        form_data["it_portal_username"] = c1.text_input("IT Portal Username", value=gs("it_portal_username"))
        form_data["it_portal_password"] = c2.text_input("IT Portal Password", value="Stored in password manager", disabled=True)

        section_hdr("📋 TDS / TAN")
        c1, c2, c3 = st.columns(3)
        form_data["tan_applicable"] = int(c1.checkbox("TAN Applicable", value=gi("tan_applicable")))
        form_data["traces_registered"] = int(c2.checkbox("TRACES Registered", value=gi("traces_registered")))
        form_data["tds_return_filing"] = int(c3.checkbox("TDS Return Filing", value=gi("tds_return_filing")))
        form_data["tan"] = st.text_input("TAN (10-char)", value=gs("tan"), max_chars=10)
        form_data["tds_forms_applicable"] = json.dumps(
            st.multiselect("TDS Forms Applicable", TDS_FORMS, default=parse_json_list("tds_forms_applicable"))
        )

        section_hdr("📄 ITR Details")
        c1, c2 = st.columns(2)
        itr_idx = ITR_FORMS.index(gs("itr_form_type", "ITR-1")) if gs("itr_form_type", "ITR-1") in ITR_FORMS else 0
        form_data["itr_form_type"] = c1.selectbox("ITR Form Type", ITR_FORMS, index=itr_idx)
        form_data["tax_audit_44ab"] = int(c2.checkbox("Tax Audit u/s 44AB", value=gi("tax_audit_44ab")))
        form_data["advance_tax_tracking"] = int(st.checkbox("Track Advance Tax", value=gi("advance_tax_tracking")))

    # ── TAB D: GST ───────────────────────────────────────────────────────────
    with tabs[3]:
        section_hdr("🧾 GST Details")
        c1, c2 = st.columns(2)
        gstin_val = c1.text_input("GSTIN", value=gs("gstin"), max_chars=15)
        form_data["gstin"] = gstin_val.upper()
        if gstin_val:
            if validate_gstin(gstin_val):
                c1.markdown('<span class="val-ok">✅ Valid GSTIN format</span>', unsafe_allow_html=True)
            else:
                c1.markdown('<span class="val-err">❌ Invalid GSTIN format</span>', unsafe_allow_html=True)
        grt_idx = GST_REG_TYPES.index(gs("gst_registration_type", "Regular")) if gs("gst_registration_type", "Regular") in GST_REG_TYPES else 0
        form_data["gst_registration_type"] = c2.selectbox("GST Registration Type", GST_REG_TYPES, index=grt_idx)

        c1, c2 = st.columns(2)
        grd = gd("gst_registration_date")
        grd_val = c1.date_input("GST Registration Date", value=grd, format="DD/MM/YYYY")
        form_data["gst_registration_date"] = grd_val.isoformat() if grd_val else ""
        form_data["gst_annual_turnover"] = c2.number_input("GST Annual Turnover (₹)", value=gf("gst_annual_turnover"), min_value=0.0, step=10000.0, format="%.2f")

        c1, c2 = st.columns(2)
        gff_idx = GSTR_FREQUENCY.index(gs("gstr_filing_frequency", "Monthly")) if gs("gstr_filing_frequency", "Monthly") in GSTR_FREQUENCY else 0
        form_data["gstr_filing_frequency"] = c1.selectbox("GSTR Filing Frequency", GSTR_FREQUENCY, index=gff_idx)
        form_data["gst_portal_username"] = c2.text_input("GST Portal Username", value=gs("gst_portal_username"))
        form_data["gst_portal_password"] = "Stored in password manager"

        form_data["gst_practitioner_auth"] = int(st.checkbox("GST Practitioner Authorised", value=gi("gst_practitioner_auth")))
        existing_gstr = parse_json_list("gst_returns_in_scope")
        form_data["gst_returns_in_scope"] = json.dumps(
            st.multiselect("GST Returns in Scope", GSTR_RETURNS, default=existing_gstr)
        )
        c1, c2 = st.columns(2)
        form_data["gstr_9c_recon"] = int(c1.checkbox("GSTR-9C Reconciliation", value=gi("gstr_9c_recon")))
        form_data["e_way_bill_filing"] = int(c2.checkbox("E-Way Bill Filing", value=gi("e_way_bill_filing")))

    # ── TAB E: ROC/MCA ───────────────────────────────────────────────────────
    with tabs[4]:
        current_const = form_data.get("constitution", gs("constitution", "Individual"))
        if current_const not in COMPANY_CONSTITUTIONS:
            st.info(f"🏢 ROC/MCA section is applicable only for: {', '.join(COMPANY_CONSTITUTIONS)}. Current constitution: **{current_const}**")
        else:
            section_hdr("🏢 ROC / MCA Details")
            c1, c2 = st.columns(2)
            form_data["cin_llpin"] = c1.text_input("CIN / LLPIN", value=gs("cin_llpin"))
            doi = gd("date_of_incorporation")
            doi_val = c2.date_input("Date of Incorporation", value=doi, format="DD/MM/YYYY")
            form_data["date_of_incorporation"] = doi_val.isoformat() if doi_val else ""

            form_data["registered_office_addr"] = st.text_area("Registered Office Address", value=gs("registered_office_addr"), height=70)

            c1, c2, c3 = st.columns(3)
            form_data["authorised_capital"] = c1.number_input("Authorised Capital (₹)", value=gf("authorised_capital"), min_value=0.0, step=100000.0, format="%.2f")
            form_data["paidup_capital"] = c2.number_input("Paid-up Capital (₹)", value=gf("paidup_capital"), min_value=0.0, step=100000.0, format="%.2f")
            form_data["number_of_directors"] = c3.number_input("No. of Directors", value=int(g("number_of_directors", 0) or 0), min_value=0)

            section_hdr("👥 Directors / Partners")
            existing_directors = parse_json_list("director_names_din")
            if not existing_directors:
                existing_directors = [{"name": "", "din": "", "email": "", "kyc_status": "Active"}]
            dir_df = pd.DataFrame(existing_directors)
            dir_df_edited = st.data_editor(
                dir_df, num_rows="dynamic", use_container_width=True,
                column_config={
                    "name": st.column_config.TextColumn("Name"),
                    "din": st.column_config.TextColumn("DIN / DPIN"),
                    "email": st.column_config.TextColumn("Email"),
                    "kyc_status": st.column_config.SelectboxColumn("KYC Status", options=["Active", "Pending", "Expired"]),
                }
            )
            form_data["director_names_din"] = json.dumps(dir_df_edited.to_dict("records"))

            c1, c2 = st.columns(2)
            form_data["mca_portal_username"] = c1.text_input("MCA Portal Username", value=gs("mca_portal_username"))
            form_data["mca_portal_password"] = "Stored in password manager"

            section_hdr("📋 Company Details")
            c1, c2 = st.columns(2)
            form_data["nature_of_business"] = c1.text_input("Nature of Business (NIC)", value=gs("nature_of_business"))
            fye_options = ["March 31", "December 31", "Other"]
            fye_val = gs("financial_year_end", "March 31")
            form_data["financial_year_end"] = c2.selectbox("Financial Year End", fye_options,
                index=fye_options.index(fye_val) if fye_val in fye_options else 0)

            c1, c2 = st.columns(2)
            agm = gd("agm_date_current_year")
            agm_val = c1.date_input("AGM Date (Current Year)", value=agm, format="DD/MM/YYYY")
            form_data["agm_date_current_year"] = agm_val.isoformat() if agm_val else ""
            aapd = gd("auditor_appt_date")
            aapd_val = c2.date_input("Auditor Appointment Date", value=aapd, format="DD/MM/YYYY")
            form_data["auditor_appt_date"] = aapd_val.isoformat() if aapd_val else ""

            c1, c2 = st.columns(2)
            form_data["auditor_tenure_yrs"] = c1.number_input("Auditor Tenure (Yrs)", value=int(g("auditor_tenure_yrs", 0) or 0), min_value=0, max_value=10)
            form_data["previous_auditor_name"] = c2.text_input("Previous Auditor Name", value=gs("previous_auditor_name"))

            c1, c2 = st.columns(2)
            form_data["caro_applicable"] = int(c1.checkbox("CARO 2020 Applicable", value=gi("caro_applicable")))
            form_data["listed_entity"] = int(c2.checkbox("Listed Entity (SEBI LODR)", value=gi("listed_entity")))

            existing_roc = parse_json_list("roc_forms_in_scope")
            form_data["roc_forms_in_scope"] = json.dumps(
                st.multiselect("ROC Forms in Scope", ROC_FORMS, default=existing_roc)
            )

    # ── TAB F: Services ──────────────────────────────────────────────────────
    with tabs[5]:
        section_hdr("⚙️ Services in Scope")
        c1, c2, c3 = st.columns(3)
        form_data["itr_filing"] = int(c1.checkbox("📄 ITR Filing", value=gi("itr_filing")))
        form_data["gst_compliance"] = int(c2.checkbox("🧾 GST Compliance", value=gi("gst_compliance")))
        form_data["statutory_audit"] = int(c3.checkbox("🔍 Statutory Audit", value=gi("statutory_audit")))

        c1, c2, c3 = st.columns(3)
        form_data["tax_audit_3cd"] = int(c1.checkbox("📋 Tax Audit (3CD)", value=gi("tax_audit_3cd")))
        form_data["roc_mca_compliance"] = int(c2.checkbox("🏛️ ROC / MCA", value=gi("roc_mca_compliance")))
        form_data["bookkeeping"] = int(c3.checkbox("📒 Bookkeeping", value=gi("bookkeeping")))

        c1, c2, c3 = st.columns(3)
        form_data["payroll_processing"] = int(c1.checkbox("💰 Payroll Processing", value=gi("payroll_processing")))
        form_data["internal_audit"] = int(c2.checkbox("🔎 Internal Audit", value=gi("internal_audit")))
        form_data["tds_return_service"] = int(c3.checkbox("📑 TDS Returns", value=gi("tds_return_service")))

        c1, c2 = st.columns(2)
        form_data["fema_rbi_advisory"] = int(c1.checkbox("🌐 FEMA / RBI Advisory", value=gi("fema_rbi_advisory")))
        form_data["valuation_services"] = int(c2.checkbox("📈 Valuation Services", value=gi("valuation_services")))

        if form_data.get("bookkeeping"):
            bf_idx = BOOKING_FREQ.index(gs("bookkeeping_frequency", "Monthly")) if gs("bookkeeping_frequency", "Monthly") in BOOKING_FREQ else 0
            form_data["bookkeeping_frequency"] = st.selectbox("Bookkeeping Frequency", BOOKING_FREQ, index=bf_idx)
        if form_data.get("internal_audit"):
            af_idx = AUDIT_FREQ.index(gs("internal_audit_freq", "Monthly")) if gs("internal_audit_freq", "Monthly") in AUDIT_FREQ else 0
            form_data["internal_audit_freq"] = st.selectbox("Internal Audit Frequency", AUDIT_FREQ, index=af_idx)

    # ── TAB G: Engagement ────────────────────────────────────────────────────
    with tabs[6]:
        section_hdr("📄 Engagement Letter & Onboarding")
        c1, c2 = st.columns(2)
        form_data["el_issued"] = int(c1.checkbox("EL Issued", value=gi("el_issued")))
        form_data["el_signed_by_client"] = int(c2.checkbox("EL Signed by Client", value=gi("el_signed_by_client")))

        c1, c2 = st.columns(2)
        eld = gd("el_issue_date")
        eld_val = c1.date_input("EL Issue Date", value=eld, format="DD/MM/YYYY")
        form_data["el_issue_date"] = eld_val.isoformat() if eld_val else ""
        elr = gd("el_last_renewed")
        elr_val = c2.date_input("EL Last Renewed", value=elr, format="DD/MM/YYYY")
        form_data["el_last_renewed"] = elr_val.isoformat() if elr_val else ""

        slu = gd("scope_last_updated")
        slu_val = st.date_input("Scope Last Updated", value=slu, format="DD/MM/YYYY")
        form_data["scope_last_updated"] = slu_val.isoformat() if slu_val else ""

        section_hdr("✅ Onboarding Checklist")
        c1, c2, c3 = st.columns(3)
        form_data["kyc_docs_collected"] = int(c1.checkbox("KYC Docs Collected", value=gi("kyc_docs_collected")))
        form_data["onboarding_complete"] = int(c2.checkbox("Onboarding Complete", value=gi("onboarding_complete")))
        form_data["poa_auth_letter"] = int(c3.checkbox("POA / Auth Letter", value=gi("poa_auth_letter")))

    # ── TAB H: Compliance ────────────────────────────────────────────────────
    with tabs[7]:
        section_hdr("📅 Last Filed — Compliance Tracker")
        c1, c2 = st.columns(2)
        form_data["last_itr_ay"] = c1.text_input("Last ITR Filed (AY)", value=gs("last_itr_ay"), placeholder="e.g. AY 2024-25")
        form_data["itr_ack_number"] = c2.text_input("ITR Acknowledgement No.", value=gs("itr_ack_number"), max_chars=15)
        c1, c2 = st.columns(2)
        form_data["last_gst_return"] = c1.text_input("Last GST Return", value=gs("last_gst_return"), placeholder="e.g. GSTR-3B Mar 2025")
        form_data["last_tds_return"] = c2.text_input("Last TDS Return", value=gs("last_tds_return"), placeholder="e.g. Q2 26Q")
        c1, c2 = st.columns(2)
        form_data["last_roc_filing"] = c1.text_input("Last ROC Filing Date", value=gs("last_roc_filing"))
        form_data["last_roc_form_filed"] = c2.text_input("Last ROC Form Filed", value=gs("last_roc_form_filed"))

        section_hdr("⏳ Pending Items")
        c1, c2 = st.columns(2)
        form_data["pending_with_client"] = c1.text_area("Pending With Client", value=gs("pending_with_client"), height=80)
        form_data["pending_with_ca"] = c2.text_area("Pending With CA", value=gs("pending_with_ca"), height=80)

        section_hdr("⚠️ Notices & Proceedings")
        form_data["it_notices_scrutiny"] = st.text_area("IT Notices / Scrutiny", value=gs("it_notices_scrutiny"), height=70)
        form_data["gst_notices"] = st.text_area("GST Notices", value=gs("gst_notices"), height=70)
        form_data["other_proceedings"] = st.text_area("Other Proceedings", value=gs("other_proceedings"), height=70)

    # ── TAB I: Fees ──────────────────────────────────────────────────────────
    with tabs[8]:
        section_hdr("💰 Fee Structure")
        c1, c2 = st.columns(2)
        form_data["itr_filing_fee"] = c1.number_input("ITR Filing Fee (₹)", value=gf("itr_filing_fee"), min_value=0.0, step=500.0, format="%.2f")
        form_data["gst_monthly_retainer"] = c2.number_input("GST Monthly Retainer (₹)", value=gf("gst_monthly_retainer"), min_value=0.0, step=500.0, format="%.2f")
        c1, c2 = st.columns(2)
        form_data["gstr9_9c_fee"] = c1.number_input("GSTR-9 / 9C Fee (₹)", value=gf("gstr9_9c_fee"), min_value=0.0, step=500.0, format="%.2f")
        form_data["tds_return_fee"] = c2.number_input("TDS Return Fee (₹)", value=gf("tds_return_fee"), min_value=0.0, step=500.0, format="%.2f")
        c1, c2 = st.columns(2)
        form_data["statutory_audit_fee"] = c1.number_input("Statutory Audit Fee (₹)", value=gf("statutory_audit_fee"), min_value=0.0, step=1000.0, format="%.2f")
        form_data["tax_audit_3cd_fee"] = c2.number_input("Tax Audit (3CD) Fee (₹)", value=gf("tax_audit_3cd_fee"), min_value=0.0, step=1000.0, format="%.2f")
        c1, c2 = st.columns(2)
        form_data["roc_mca_annual_fee"] = c1.number_input("ROC/MCA Annual Fee (₹)", value=gf("roc_mca_annual_fee"), min_value=0.0, step=1000.0, format="%.2f")
        form_data["bookkeeping_monthly"] = c2.number_input("Bookkeeping Monthly (₹)", value=gf("bookkeeping_monthly"), min_value=0.0, step=500.0, format="%.2f")
        c1, c2 = st.columns(2)
        form_data["payroll_fee"] = c1.number_input("Payroll Fee (₹)", value=gf("payroll_fee"), min_value=0.0, step=500.0, format="%.2f")
        form_data["other_services_fee"] = c2.number_input("Other Services Fee (₹)", value=gf("other_services_fee"), min_value=0.0, step=500.0, format="%.2f")

        total = calc_total_fee(form_data)
        form_data["total_annual_fee"] = total
        st.metric("📊 Total Annual Fee (₹)", f"₹ {total:,.2f}")
        form_data["gst_on_fees"] = int(st.checkbox("GST Applicable on Fees (>₹20L)", value=gi("gst_on_fees")))

        section_hdr("🧾 Billing & Collections (FY)")
        c1, c2 = st.columns(2)
        pt_idx = PAYMENT_TERMS.index(gs("payment_terms", "Monthly")) if gs("payment_terms", "Monthly") in PAYMENT_TERMS else 0
        form_data["payment_terms"] = c1.selectbox("Payment Terms", PAYMENT_TERMS, index=pt_idx)
        if_idx = INVOICE_FREQ.index(gs("invoice_frequency", "Monthly")) if gs("invoice_frequency", "Monthly") in INVOICE_FREQ else 0
        form_data["invoice_frequency"] = c2.selectbox("Invoice Frequency", INVOICE_FREQ, index=if_idx)

        c1, c2 = st.columns(2)
        lid = gd("last_invoice_date")
        lid_val = c1.date_input("Last Invoice Date", value=lid, format="DD/MM/YYYY")
        form_data["last_invoice_date"] = lid_val.isoformat() if lid_val else ""
        form_data["last_invoice_amount"] = c2.number_input("Last Invoice Amount (₹)", value=gf("last_invoice_amount"), min_value=0.0, step=100.0, format="%.2f")

        c1, c2 = st.columns(2)
        form_data["total_billed_fy"] = c1.number_input("Total Billed FY (₹)", value=gf("total_billed_fy"), min_value=0.0, step=1000.0, format="%.2f")
        form_data["total_received_fy"] = c2.number_input("Total Received FY (₹)", value=gf("total_received_fy"), min_value=0.0, step=1000.0, format="%.2f")

        outstanding = form_data["total_billed_fy"] - form_data["total_received_fy"]
        form_data["outstanding_balance"] = outstanding
        col = "🔴" if outstanding > 0 else "🟢"
        st.metric(f"{col} Outstanding Balance (₹)", f"₹ {outstanding:,.2f}")

        frd = gd("fee_revision_due")
        frd_val = st.date_input("Fee Revision Due Date", value=frd, format="DD/MM/YYYY")
        form_data["fee_revision_due"] = frd_val.isoformat() if frd_val else ""
        form_data["fee_notes"] = st.text_area("Fee Notes", value=gs("fee_notes"), height=80)

    # ── TAB J: Bank ──────────────────────────────────────────────────────────
    with tabs[9]:
        section_hdr("🏦 Primary Bank Account")
        c1, c2 = st.columns(2)
        form_data["bank_name"] = c1.text_input("Bank Name", value=gs("bank_name"))
        form_data["account_number"] = c2.text_input("Account Number", value=gs("account_number"))
        c1, c2 = st.columns(2)
        ifsc = c1.text_input("IFSC Code", value=gs("ifsc_code"), max_chars=11)
        form_data["ifsc_code"] = ifsc.upper()
        if ifsc and not validate_ifsc(ifsc):
            c1.markdown('<span class="val-err">❌ Format: ABCD0123456</span>', unsafe_allow_html=True)
        at_idx = ACCOUNT_TYPES.index(gs("account_type", "Savings")) if gs("account_type", "Savings") in ACCOUNT_TYPES else 0
        form_data["account_type"] = c2.selectbox("Account Type", ACCOUNT_TYPES, index=at_idx)
        form_data["account_holder_name"] = st.text_input("Account Holder Name", value=gs("account_holder_name"))
        form_data["bank_linked_to_pan"] = int(st.checkbox("Bank Account Pre-validated with PAN", value=gi("bank_linked_to_pan")))

        section_hdr("🏦 Additional Bank Account")
        form_data["no_of_additional_accts"] = st.number_input("No. of Additional Accounts", value=int(g("no_of_additional_accts", 0) or 0), min_value=0)
        if form_data["no_of_additional_accts"] > 0:
            c1, c2, c3 = st.columns(3)
            form_data["bank2_name"] = c1.text_input("Bank 2 Name", value=gs("bank2_name"))
            form_data["bank2_acc_no"] = c2.text_input("Bank 2 Acc. No.", value=gs("bank2_acc_no"))
            form_data["bank2_ifsc"] = c3.text_input("Bank 2 IFSC", value=gs("bank2_ifsc")).upper()

    # ── TAB K: Registrations ─────────────────────────────────────────────────
    with tabs[10]:
        section_hdr("🔖 Other Registrations")
        c1, c2 = st.columns(2)
        form_data["pf_account_number"] = c1.text_input("PF Account Number", value=gs("pf_account_number"))
        form_data["pf_establishment_code"] = c2.text_input("PF Establishment Code", value=gs("pf_establishment_code"))
        c1, c2 = st.columns(2)
        form_data["esi_number"] = c1.text_input("ESI Number", value=gs("esi_number"))
        form_data["msme_udyam_reg_no"] = c2.text_input("MSME / Udyam Reg. No.", value=gs("msme_udyam_reg_no"))
        form_data["import_export_code"] = st.text_input("Import Export Code (IEC)", value=gs("import_export_code"))

    # ── TAB L: Risk & CRM ────────────────────────────────────────────────────
    with tabs[11]:
        section_hdr("⚠️ Risk Assessment")
        c1, c2 = st.columns(2)
        rf_idx = RISK_FLAGS.index(gs("risk_flag", "Low")) if gs("risk_flag", "Low") in RISK_FLAGS else 0
        form_data["risk_flag"] = c1.selectbox("Risk Flag", RISK_FLAGS, index=rf_idx)
        pb_idx = PAYMENT_BEHAVIOUR.index(gs("payment_behaviour", "Prompt")) if gs("payment_behaviour", "Prompt") in PAYMENT_BEHAVIOUR else 0
        form_data["payment_behaviour"] = c2.selectbox("Payment Behaviour", PAYMENT_BEHAVIOUR, index=pb_idx)

        section_hdr("🤝 Relationship")
        c1, c2 = st.columns(2)
        form_data["referred_by"] = c1.text_input("Referred By", value=gs("referred_by"))
        form_data["clients_referred"] = c2.text_input("Clients Referred (by this client)", value=gs("clients_referred"))
        ci_idx = CLIENT_IMPORTANCE.index(gs("client_importance", "Regular")) if gs("client_importance", "Regular") in CLIENT_IMPORTANCE else 0
        form_data["client_importance"] = st.selectbox("Client Importance", CLIENT_IMPORTANCE, index=ci_idx)

        section_hdr("📆 Follow-up & Review")
        c1, c2 = st.columns(2)
        nrd = gd("next_review_date")
        nrd_val = c1.date_input("Next Review Date", value=nrd, format="DD/MM/YYYY")
        form_data["next_review_date"] = nrd_val.isoformat() if nrd_val else ""
        lc = gd("last_contacted")
        lc_val = c2.date_input("Last Contacted", value=lc, format="DD/MM/YYYY")
        form_data["last_contacted"] = lc_val.isoformat() if lc_val else ""
        c1, c2 = st.columns(2)
        nfd = gd("next_followup_date")
        nfd_val = c1.date_input("Next Follow-up Date", value=nfd, format="DD/MM/YYYY")
        form_data["next_followup_date"] = nfd_val.isoformat() if nfd_val else ""
        form_data["followup_purpose"] = c2.text_input("Follow-up Purpose", value=gs("followup_purpose"))

    # ── TAB M: Internal ──────────────────────────────────────────────────────
    with tabs[12]:
        section_hdr("🗂️ Internal Reference")
        dfl = gs("drive_folder_link")
        form_data["drive_folder_link"] = st.text_input("Google Drive Folder Link", value=dfl,
            placeholder="https://drive.google.com/drive/folders/...")
        if form_data["drive_folder_link"] and not form_data["drive_folder_link"].startswith("https://drive.google.com"):
            st.markdown('<span class="val-err">❌ Must be a valid Google Drive URL</span>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        form_data["compuoffice_code"] = c1.text_input("CompuOffice Client Code", value=gs("compuoffice_code"))
        form_data["tallyprime_name"] = c2.text_input("TallyPrime Company Name", value=gs("tallyprime_name"))
        form_data["internal_notes"] = st.text_area("Internal Notes", value=gs("internal_notes"), height=150)
        if gs("client_status") == "Discontinued" or form_data.get("client_status") == "Discontinued":
            form_data["exit_discontinue_reason"] = st.text_area("Exit / Discontinue Reason", value=gs("exit_discontinue_reason"), height=80)

    return form_data


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT HELPER
# ─────────────────────────────────────────────────────────────────────────────
def export_to_excel(df_full: pd.DataFrame, summary_df: pd.DataFrame = None) -> bytes:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        pass

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet 1: Summary
        sv = summary_df if summary_df is not None else df_full
        sv.to_excel(writer, sheet_name="Client Summary", index=False)

        # Additional sheets by section column groups
        sections = {
            "Identity": ["client_id","client_full_name","constitution","date_of_birth_incorp","father_promoter_name","city","state","pincode","client_status","client_since","source_of_client"],
            "Contact": ["client_id","client_full_name","primary_mobile","alternate_mobile","primary_email","alternate_email","whatsapp_number","contact_person_name","contact_person_desig","contact_person_mobile","contact_person_email"],
            "Income Tax": ["client_id","client_full_name","pan","residential_status","tax_regime","itr_form_type","tax_audit_44ab","tan_applicable","tan"],
            "GST": ["client_id","client_full_name","gstin","gst_registration_type","gst_annual_turnover","gstr_filing_frequency","gst_practitioner_auth"],
            "Fees": ["client_id","client_full_name","total_annual_fee","total_billed_fy","total_received_fy","outstanding_balance","payment_terms","fee_notes"],
            "Risk & CRM": ["client_id","client_full_name","risk_flag","payment_behaviour","client_importance","next_review_date","last_contacted","next_followup_date"],
        }
        for sheet_name, cols in sections.items():
            valid_cols = [c for c in cols if c in df_full.columns]
            df_full[valid_cols].to_excel(writer, sheet_name=sheet_name, index=False)

        # Style sheets
        try:
            for ws in writer.book.worksheets:
                hdr_fill = PatternFill("solid", fgColor="1A2A3A")
                hdr_font = Font(color="E8B86D", bold=True, size=10)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                for col_idx, col in enumerate(ws.columns, 1):
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
        except Exception:
            pass

    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PAGES
# ─────────────────────────────────────────────────────────────────────────────
def page_summary():
    st.markdown('<div class="page-title">📋 Client Summary</div>', unsafe_allow_html=True)
    df = get_summary_view()

    # KPI row
    total = len(df)
    active = len(df[df["client_status"] == "Active"])
    total_fee = df["total_annual_fee"].sum()
    outstanding = df["outstanding_balance"].sum()
    high_risk = len(df[df["risk_flag"] == "High"])

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("👥 Total Clients", total)
    k2.metric("✅ Active", active)
    k3.metric("💰 Annual Fee (₹)", f"₹{total_fee:,.0f}")
    k4.metric("⚠️ Outstanding (₹)", f"₹{outstanding:,.0f}")
    k5.metric("🔴 High Risk", high_risk)

    st.markdown("---")

    # Quick filter chips
    today = date.today()
    review_30 = (pd.to_datetime(df["next_review_date"], errors="coerce").dt.date <= today + timedelta(days=30)) & \
                (pd.to_datetime(df["next_review_date"], errors="coerce").dt.date >= today)

    filter_labels = ["All", "Active", "Prospect", "High Risk", "EL Pending", "Review Due (30d)"]
    selected_chip = st.radio("Quick Filter:", filter_labels, horizontal=True, label_visibility="collapsed")

    if selected_chip == "Active":
        df = df[df["client_status"] == "Active"]
    elif selected_chip == "Prospect":
        df = df[df["client_status"] == "Prospect"]
    elif selected_chip == "High Risk":
        df = df[df["risk_flag"] == "High"]
    elif selected_chip == "EL Pending":
        df = df[df["el_signed_by_client"] == 0]
    elif selected_chip == "Review Due (30d)":
        df = df[review_30]

    st.caption(f"Showing {len(df)} client(s)")

    # Rename for display
    display_df = df.rename(columns={
        "client_id": "ID", "client_full_name": "Client Name",
        "constitution": "Constitution", "pan": "PAN", "gstin": "GSTIN",
        "primary_mobile": "Mobile", "primary_email": "Email",
        "itr_filing": "ITR", "gst_compliance": "GST",
        "statutory_audit": "Audit", "roc_mca_compliance": "ROC",
        "bookkeeping": "Books", "total_annual_fee": "Annual Fee (₹)",
        "el_signed_by_client": "EL Signed", "last_itr_ay": "Last ITR AY",
        "last_gst_return": "Last GST", "last_roc_filing": "Last ROC",
        "outstanding_balance": "Outstanding (₹)",
        "client_status": "Status", "risk_flag": "Risk",
        "next_review_date": "Next Review", "drive_folder_link": "Drive",
    })

    bool_cols = ["ITR", "GST", "Audit", "ROC", "Books", "EL Signed"]
    for bc in bool_cols:
        if bc in display_df.columns:
            display_df[bc] = display_df[bc].apply(lambda x: "✅" if x else "—")

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Annual Fee (₹)": st.column_config.NumberColumn(format="₹%.0f"),
            "Outstanding (₹)": st.column_config.NumberColumn(format="₹%.0f"),
            "Drive": st.column_config.LinkColumn("Drive 🔗"),
        }
    )


def page_add_client():
    st.markdown('<div class="page-title">➕ Add New Client</div>', unsafe_allow_html=True)

    with st.form("add_form", clear_on_submit=False):
        form_data = render_form(mode="add")
        submitted = st.form_submit_button("💾 Save Client", use_container_width=True, type="primary")

    if submitted:
        errors = []
        if not form_data.get("client_full_name", "").strip():
            errors.append("Client Full Name is required.")
        if not form_data.get("constitution"):
            errors.append("Constitution is required.")
        if not form_data.get("primary_mobile", "").strip():
            errors.append("Primary Mobile is required.")
        elif not validate_mobile(form_data["primary_mobile"]):
            errors.append("Primary Mobile must be a valid 10-digit number.")
        if not form_data.get("primary_email", "").strip():
            errors.append("Primary Email is required.")
        if form_data.get("pan") and not validate_pan(form_data["pan"]):
            errors.append("PAN format is invalid.")
        if form_data.get("gstin") and not validate_gstin(form_data["gstin"]):
            errors.append("GSTIN format is invalid.")

        if errors:
            for e in errors:
                st.error(f"❌ {e}")
        else:
            # Mask aadhaar
            if form_data.get("aadhaar_number"):
                form_data["aadhaar_number"] = mask_aadhaar(form_data["aadhaar_number"])
            try:
                new_id = insert_client(form_data)
                get_all_clients.clear()
                get_summary_view.clear()
                st.success(f"✅ Client saved successfully! Client ID: **{new_id}**")
                st.balloons()
            except Exception as ex:
                st.error(f"❌ Database error: {ex}")


def page_edit_client():
    st.markdown('<div class="page-title">✏️ Edit Client</div>', unsafe_allow_html=True)

    df = get_all_clients()
    if df.empty:
        st.info("No clients found. Add a client first.")
        return

    options = df.apply(lambda r: f"{r['client_id']} — {r['client_full_name']}", axis=1).tolist()
    selected = st.selectbox("Select Client to Edit", options)
    client_id = selected.split(" — ")[0]

    client_data = get_client_by_id(client_id)
    if not client_data:
        st.error("Client not found.")
        return

    with st.form("edit_form"):
        form_data = render_form(defaults=client_data, mode="edit")
        form_data["client_id"] = client_id
        col1, col2 = st.columns([4, 1])
        save_btn = col1.form_submit_button("💾 Update Client", use_container_width=True, type="primary")
        confirm_delete = col2.form_submit_button("🗑️ Delete", use_container_width=True)

    if save_btn:
        if form_data.get("aadhaar_number") and not form_data["aadhaar_number"].startswith("X"):
            form_data["aadhaar_number"] = mask_aadhaar(form_data["aadhaar_number"])
        try:
            update_client(form_data)
            get_all_clients.clear()
            get_summary_view.clear()
            st.success(f"✅ Client **{client_id}** updated successfully!")
        except Exception as ex:
            st.error(f"❌ Error: {ex}")

    if confirm_delete:
        if "delete_confirm" not in st.session_state:
            st.session_state["delete_confirm"] = False
        st.session_state["delete_confirm"] = True

    if st.session_state.get("delete_confirm"):
        st.warning(f"⚠️ You are about to permanently delete **{client_data.get('client_full_name')}**.")
        col1, col2 = st.columns(2)
        if col1.button("🔴 Confirm Delete", use_container_width=True):
            delete_client(client_id)
            get_all_clients.clear()
            get_summary_view.clear()
            st.session_state["delete_confirm"] = False
            st.success("Client deleted.")
            st.rerun()
        if col2.button("Cancel", use_container_width=True):
            st.session_state["delete_confirm"] = False
            st.rerun()


def page_search():
    st.markdown('<div class="page-title">🔍 Search & Filter</div>', unsafe_allow_html=True)

    if "filters" not in st.session_state:
        st.session_state["filters"] = {}

    with st.sidebar:
        st.markdown("## 🔍 Filters")
        filters = {}
        filters["text"] = st.text_input("🔎 Search (Name / PAN / GSTIN / Mobile / Email)")
        filters["constitution"] = st.multiselect("Constitution", CONSTITUTIONS)
        filters["client_status"] = st.multiselect("Client Status", CLIENT_STATUSES)
        filters["risk_flag"] = st.multiselect("Risk Flag", RISK_FLAGS)
        filters["client_importance"] = st.multiselect("Client Importance", CLIENT_IMPORTANCE)
        st.markdown("**Services**")
        filters["itr_filing"] = st.checkbox("ITR Filing")
        filters["gst_compliance"] = st.checkbox("GST Compliance")
        filters["statutory_audit"] = st.checkbox("Statutory Audit")
        filters["roc_mca_compliance"] = st.checkbox("ROC / MCA")
        filters["bookkeeping"] = st.checkbox("Bookkeeping")
        filters["outstanding_only"] = st.toggle("Outstanding > 0")
        filters["el_signed"] = st.radio("EL Signed", ["All", "Yes", "No"])
        filters["state"] = st.multiselect("State", INDIAN_STATES)
        filters["city"] = st.text_input("City")
        c1, c2 = st.columns(2)
        filters["fee_min"] = c1.number_input("Fee Min (₹)", min_value=0.0, step=1000.0)
        filters["fee_max"] = c2.number_input("Fee Max (₹)", min_value=0.0, step=1000.0, value=10000000.0)
        if st.button("🔄 Clear All Filters"):
            st.rerun()

    st.session_state["filters"] = filters
    results = search_clients(filters)
    st.caption(f"**{len(results)}** client(s) match your filters")

    if results.empty:
        st.info("No clients match the current filters.")
        return

    display_df = results[["client_id","client_full_name","constitution","pan","gstin",
                           "primary_mobile","primary_email","client_status","risk_flag",
                           "total_annual_fee","outstanding_balance","next_review_date"]].rename(columns={
        "client_id": "ID", "client_full_name": "Name", "constitution": "Constitution",
        "pan": "PAN", "gstin": "GSTIN", "primary_mobile": "Mobile",
        "primary_email": "Email", "client_status": "Status", "risk_flag": "Risk",
        "total_annual_fee": "Annual Fee (₹)", "outstanding_balance": "Outstanding (₹)",
        "next_review_date": "Next Review",
    })
    st.dataframe(display_df, use_container_width=True, hide_index=True,
        column_config={
            "Annual Fee (₹)": st.column_config.NumberColumn(format="₹%.0f"),
            "Outstanding (₹)": st.column_config.NumberColumn(format="₹%.0f"),
        }
    )


def page_analytics():
    try:
        import plotly.express as px
        import plotly.graph_objects as go
    except ImportError:
        st.error("Install plotly: pip install plotly")
        return

    st.markdown('<div class="page-title">📊 Analytics Dashboard</div>', unsafe_allow_html=True)
    df = get_all_clients()
    if df.empty:
        st.info("No data yet.")
        return

    # KPI row
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("👥 Total Clients", len(df))
    k2.metric("✅ Active", len(df[df["client_status"] == "Active"]))
    k3.metric("💰 Annual Fee (₹)", f"₹{df['total_annual_fee'].sum():,.0f}")
    k4.metric("⚠️ Outstanding (₹)", f"₹{df['outstanding_balance'].sum():,.0f}")
    k5.metric("🔴 High Risk", len(df[df["risk_flag"] == "High"]))
    st.markdown("---")

    # Row 2: Constitution bar + Status pie
    c1, c2 = st.columns(2)
    with c1:
        const_counts = df["constitution"].value_counts().reset_index()
        const_counts.columns = ["Constitution", "Count"]
        fig1 = px.bar(const_counts, x="Count", y="Constitution", orientation="h",
                      title="Clients by Constitution", color_discrete_sequence=["#e8b86d"])
        fig1.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                           title_font=dict(family="Playfair Display", size=16))
        st.plotly_chart(fig1, use_container_width=True)
    with c2:
        status_counts = df["client_status"].value_counts().reset_index()
        status_counts.columns = ["Status", "Count"]
        fig2 = px.pie(status_counts, names="Status", values="Count",
                      title="Clients by Status", hole=0.4,
                      color_discrete_sequence=px.colors.sequential.Blues_r)
        fig2.update_layout(title_font=dict(family="Playfair Display", size=16))
        st.plotly_chart(fig2, use_container_width=True)

    # Row 3: Services heatmap + Fee by constitution
    c1, c2 = st.columns(2)
    with c1:
        svc_cols = {"itr_filing": "ITR Filing", "gst_compliance": "GST Compliance",
                    "statutory_audit": "Statutory Audit", "tax_audit_3cd": "Tax Audit (3CD)",
                    "roc_mca_compliance": "ROC/MCA", "bookkeeping": "Bookkeeping",
                    "payroll_processing": "Payroll", "tds_return_service": "TDS Returns",
                    "internal_audit": "Internal Audit"}
        svc_data = [(label, int(df[col].sum())) for col, label in svc_cols.items() if col in df.columns]
        svc_df = pd.DataFrame(svc_data, columns=["Service", "Clients"]).sort_values("Clients")
        fig3 = px.bar(svc_df, x="Clients", y="Service", orientation="h",
                      title="Services Adoption", color="Clients",
                      color_continuous_scale=["#d4e6f1", "#1a2a3a"])
        fig3.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                           title_font=dict(family="Playfair Display", size=16))
        st.plotly_chart(fig3, use_container_width=True)
    with c2:
        fee_by_const = df.groupby("constitution")["total_annual_fee"].sum().reset_index()
        fee_by_const.columns = ["Constitution", "Total Fee (₹)"]
        fig4 = px.bar(fee_by_const, x="Constitution", y="Total Fee (₹)",
                      title="Total Annual Fee by Constitution",
                      color_discrete_sequence=["#1a2a3a"])
        fig4.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                           title_font=dict(family="Playfair Display", size=16))
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown("---")
    st.subheader("📋 Action Tables")

    today = date.today()

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**🏆 Top 10 Clients by Annual Fee**")
        top10 = df.nlargest(10, "total_annual_fee")[["client_full_name","constitution","total_annual_fee"]].rename(
            columns={"client_full_name": "Client", "constitution": "Type", "total_annual_fee": "Annual Fee (₹)"})
        st.dataframe(top10, use_container_width=True, hide_index=True,
            column_config={"Annual Fee (₹)": st.column_config.NumberColumn(format="₹%.0f")})

    with col2:
        st.markdown("**⚠️ Outstanding Dues**")
        dues = df[df["outstanding_balance"] > 0][["client_full_name","outstanding_balance","payment_behaviour"]].sort_values("outstanding_balance", ascending=False)
        dues = dues.rename(columns={"client_full_name": "Client", "outstanding_balance": "Outstanding (₹)", "payment_behaviour": "Pay Behaviour"})
        st.dataframe(dues, use_container_width=True, hide_index=True,
            column_config={"Outstanding (₹)": st.column_config.NumberColumn(format="₹%.0f")})

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**📅 Reviews Due Within 30 Days**")
        review_df = df.copy()
        review_df["nrd"] = pd.to_datetime(review_df["next_review_date"], errors="coerce").dt.date
        review_due = review_df[(review_df["nrd"] >= today) & (review_df["nrd"] <= today + timedelta(days=30))][["client_full_name","next_review_date","client_status"]].rename(
            columns={"client_full_name": "Client", "next_review_date": "Review Date", "client_status": "Status"})
        if review_due.empty:
            st.success("✅ No reviews due in next 30 days")
        else:
            st.dataframe(review_due, use_container_width=True, hide_index=True)

    with col2:
        st.markdown("**📄 EL Not Signed**")
        el_pending = df[df["el_signed_by_client"] == 0][["client_full_name","client_status","el_issue_date"]].rename(
            columns={"client_full_name": "Client", "client_status": "Status", "el_issue_date": "EL Date"})
        if el_pending.empty:
            st.success("✅ All ELs signed")
        else:
            st.dataframe(el_pending, use_container_width=True, hide_index=True)

    st.markdown("**📞 Not Contacted in 60+ Days**")
    nc_df = df.copy()
    nc_df["lc"] = pd.to_datetime(nc_df["last_contacted"], errors="coerce").dt.date
    not_contacted = nc_df[(nc_df["lc"].notna()) & (nc_df["lc"] < today - timedelta(days=60))][["client_full_name","last_contacted","primary_mobile","primary_email"]].rename(
        columns={"client_full_name": "Client", "last_contacted": "Last Contact", "primary_mobile": "Mobile", "primary_email": "Email"})
    if not_contacted.empty:
        st.success("✅ All clients contacted recently")
    else:
        st.dataframe(not_contacted, use_container_width=True, hide_index=True)


def page_export():
    st.markdown('<div class="page-title">⬇️ Export Data</div>', unsafe_allow_html=True)
    df_full = get_all_clients()
    summary_df = get_summary_view()
    today_str = date.today().strftime("%Y%m%d")

    st.markdown("### 📦 Export Full Database")
    st.write(f"Exports all **{len(df_full)}** clients across multiple sheets (Identity, Contact, Tax, GST, Fees, Risk).")
    if st.button("Generate Full Export", use_container_width=True):
        with st.spinner("Preparing Excel file..."):
            excel_bytes = export_to_excel(df_full, summary_df)
        st.download_button(
            label="⬇️ Download Full Export (.xlsx)",
            data=excel_bytes,
            file_name=f"ca_clients_full_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("### 🔍 Export Filtered View")
    filters = st.session_state.get("filters", {})
    if filters and any(v for v in filters.values()):
        filtered_df = search_clients(filters)
        st.write(f"Exports **{len(filtered_df)}** clients matching current Search & Filter criteria.")
    else:
        filtered_df = df_full
        st.info("No active filters. Will export all clients. Apply filters on the 🔍 Search page first.")

    if st.button("Generate Filtered Export", use_container_width=True):
        with st.spinner("Preparing Excel file..."):
            excel_bytes = export_to_excel(filtered_df)
        st.download_button(
            label="⬇️ Download Filtered Export (.xlsx)",
            data=excel_bytes,
            file_name=f"ca_clients_filtered_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("### 💾 Database Backup")
    st.info(f"Your SQLite database is stored at: `{os.path.abspath(DB_PATH)}`\n\n"
            "To backup, simply copy this `.db` file to a safe location. No other steps needed.")


# ─────────────────────────────────────────────────────────────────────────────
# AUTHENTICATION  (zero external dependencies — hashlib + secrets + json only)
# ─────────────────────────────────────────────────────────────────────────────

def _hash_password(password: str, salt: str = None):
    """PBKDF2-HMAC-SHA256 hash. Returns (salt, hash) both as hex strings."""
    if salt is None:
        salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 260_000)
    return salt, key.hex()


def _verify_password(password: str, salt: str, stored_hash: str) -> bool:
    _, computed = _hash_password(password, salt)
    return secrets.compare_digest(computed, stored_hash)


def _load_users() -> dict:
    """Load users.json; auto-create with default admin on first run."""
    if not os.path.exists(AUTH_YAML):
        salt, pw_hash = _hash_password("admin@123")
        default = {
            "admin": {
                "name": "Administrator",
                "email": "admin@ca.com",
                "salt": salt,
                "password": pw_hash,
                "role": "admin",
            }
        }
        with open(AUTH_YAML, "w") as f:
            json.dump(default, f, indent=2)
    with open(AUTH_YAML, "r") as f:
        return json.load(f)


def _save_users(users: dict):
    with open(AUTH_YAML, "w") as f:
        json.dump(users, f, indent=2)


def _login_check(username: str, password: str, users: dict) -> bool:
    u = users.get(username)
    if not u:
        return False
    return _verify_password(password, u["salt"], u["password"])


# ── Login page CSS ────────────────────────────────────────────────────────────
def inject_login_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700&family=Lato:wght@300;400;700&display=swap');
    html, body, [class*="css"] { font-family: 'Lato', sans-serif; }
    .stApp { background: linear-gradient(135deg, #0f1923 0%, #1a2a3a 50%, #0f1923 100%) !important; }
    .login-wrap { max-width: 420px; margin: 6vh auto 0; padding: 40px 44px;
        background: rgba(255,255,255,0.04); border: 1px solid rgba(232,184,109,0.25);
        border-radius: 12px; backdrop-filter: blur(12px);
        box-shadow: 0 24px 64px rgba(0,0,0,0.5); }
    .login-logo { text-align:center; font-family:'Playfair Display',serif;
        font-size:2.4rem; font-weight:700; color:#e8b86d; letter-spacing:0.04em; margin-bottom:4px; }
    .login-sub  { text-align:center; color:#7a9bb5; font-size:0.82rem;
        letter-spacing:0.12em; text-transform:uppercase; margin-bottom:28px; }
    .login-hint { text-align:center; color:#4a6a80; font-size:0.74rem;
        margin-top:18px; border-top:1px solid rgba(255,255,255,0.06); padding-top:14px; }
    [data-testid="stForm"] { background:transparent !important; border:none !important; }
    .stTextInput label { color:#a8c4d8 !important; font-size:0.8rem;
        letter-spacing:0.08em; text-transform:uppercase; }
    .stTextInput input { background:rgba(255,255,255,0.07) !important;
        border:1px solid rgba(232,184,109,0.3) !important; color:#e8ecf0 !important;
        border-radius:6px !important; }
    .stTextInput input:focus { border-color:#e8b86d !important;
        box-shadow:0 0 0 2px rgba(232,184,109,0.15) !important; }
    .stButton > button { width:100%;
        background:linear-gradient(90deg,#e8b86d,#d4a054) !important;
        color:#1a2a3a !important; font-weight:700 !important; font-size:0.95rem !important;
        border:none !important; border-radius:6px !important; padding:12px 0 !important;
        letter-spacing:0.06em !important; margin-top:6px; }
    .stButton > button:hover { filter:brightness(1.08);
        box-shadow:0 6px 20px rgba(232,184,109,0.3) !important; }
    </style>
    """, unsafe_allow_html=True)


# ── Login page renderer ───────────────────────────────────────────────────────
def show_login_page(users: dict):
    inject_login_css()
    st.markdown('<div class="login-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="login-logo">⚖️ CA Client Master</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-sub">Chartered Accountant Practice Manager</div>', unsafe_allow_html=True)

    with st.form("login_form", clear_on_submit=False):
        uname = st.text_input("Username")
        pwd   = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign In", use_container_width=True)

    if submitted:
        if _login_check(uname.strip(), pwd, users):
            u = users[uname.strip()]
            st.session_state["auth_ok"]   = True
            st.session_state["auth_user"] = uname.strip()
            st.session_state["auth_name"] = u["name"]
            st.session_state["auth_role"] = u.get("role", "user")
            st.rerun()
        else:
            st.error("❌ Incorrect username or password.")

    st.markdown('<div class="login-hint">🔒 All data stored locally — no cloud sync</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ── User Management panel (admin only) ───────────────────────────────────────
def show_user_management(users: dict):
    st.markdown('<div class="page-title">👥 User Management</div>', unsafe_allow_html=True)

    # Current users table
    st.subheader("Current Users")
    rows = [{"Username": u, "Name": v.get("name",""), "Email": v.get("email",""), "Role": v.get("role","user")}
            for u, v in users.items()]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.markdown("---")

    tab_add, tab_pw, tab_del = st.tabs(["➕ Add User", "🔑 Change Password", "🗑️ Delete User"])

    # ── Add User ──────────────────────────────────────────────────────────────
    with tab_add:
        with st.form("add_user_form"):
            c1, c2 = st.columns(2)
            new_uname = c1.text_input("Username *", placeholder="e.g. staff1")
            new_name  = c2.text_input("Full Name *", placeholder="e.g. Priya Shah")
            c1, c2 = st.columns(2)
            new_email = c1.text_input("Email", placeholder="staff1@yourca.com")
            new_role  = c2.selectbox("Role", ["user", "admin"])
            new_pw    = st.text_input("Password *", type="password", placeholder="Min 6 characters")
            new_pw2   = st.text_input("Confirm Password *", type="password")
            if st.form_submit_button("➕ Create User", use_container_width=True, type="primary"):
                if not new_uname or not new_name or not new_pw:
                    st.error("Username, Full Name and Password are required.")
                elif new_pw != new_pw2:
                    st.error("Passwords do not match.")
                elif len(new_pw) < 6:
                    st.error("Password must be at least 6 characters.")
                elif new_uname in users:
                    st.error(f"Username '{new_uname}' already exists.")
                else:
                    salt, pw_hash = _hash_password(new_pw)
                    users[new_uname] = {"name": new_name, "email": new_email,
                                        "salt": salt, "password": pw_hash, "role": new_role}
                    _save_users(users)
                    st.success(f"✅ User **{new_uname}** created.")
                    st.rerun()

    # ── Change Password ───────────────────────────────────────────────────────
    with tab_pw:
        with st.form("change_pw_form"):
            target = st.selectbox("Select User", list(users.keys()))
            pw1 = st.text_input("New Password *", type="password")
            pw2 = st.text_input("Confirm New Password *", type="password")
            if st.form_submit_button("🔑 Update Password", use_container_width=True, type="primary"):
                if not pw1:
                    st.error("Password cannot be empty.")
                elif pw1 != pw2:
                    st.error("Passwords do not match.")
                elif len(pw1) < 6:
                    st.error("Password must be at least 6 characters.")
                else:
                    salt, pw_hash = _hash_password(pw1)
                    users[target]["salt"]     = salt
                    users[target]["password"] = pw_hash
                    _save_users(users)
                    st.success(f"✅ Password updated for **{target}**.")

    # ── Delete User ───────────────────────────────────────────────────────────
    with tab_del:
        me = st.session_state.get("auth_user", "")
        deletable = [u for u in users if u != me]
        if not deletable:
            st.info("No other users to delete.")
        else:
            with st.form("del_user_form"):
                del_user = st.selectbox("Select User to Delete", deletable)
                confirm  = st.checkbox(f"I confirm I want to permanently delete '{del_user}'")
                if st.form_submit_button("🗑️ Delete User", use_container_width=True):
                    if not confirm:
                        st.warning("Please tick the confirmation checkbox.")
                    else:
                        del users[del_user]
                        _save_users(users)
                        st.success(f"✅ User **{del_user}** deleted.")
                        st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    inject_css()
    init_db()
    seed_sample_data()

    if "delete_confirm" not in st.session_state:
        st.session_state["delete_confirm"] = False

    with st.sidebar:
        st.markdown("## ⚖️ CA Client Master")
        st.markdown("---")
        page = st.radio(
            "Navigate",
            ["📋 Client Summary", "➕ Add New Client", "✏️ Edit Client",
             "🔍 Search & Filter", "📊 Analytics", "⬇️ Export"],
            label_visibility="collapsed"
        )
        st.markdown("---")
        df_count = get_all_clients()
        st.markdown(f"**Total Clients:** {len(df_count)}")
        active_count = len(df_count[df_count["client_status"] == "Active"]) if not df_count.empty else 0
        st.markdown(f"**Active:** {active_count}")

    if page == "📋 Client Summary":
        page_summary()
    elif page == "➕ Add New Client":
        page_add_client()
    elif page == "✏️ Edit Client":
        page_edit_client()
    elif page == "🔍 Search & Filter":
        page_search()
    elif page == "📊 Analytics":
        page_analytics()
    elif page == "⬇️ Export":
        page_export()


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
users = _load_users()

if not st.session_state.get("auth_ok"):
    show_login_page(users)
    st.stop()

# ── Authenticated ─────────────────────────────────────────────────────────────
auth_name = st.session_state.get("auth_name", "User")
auth_user = st.session_state.get("auth_user", "")
auth_role = st.session_state.get("auth_role", "user")

# Sidebar user card + sign-out (renders before main()'s sidebar block)
with st.sidebar:
    st.markdown(f"""
    <div style='background:rgba(232,184,109,0.1);border:1px solid rgba(232,184,109,0.3);
    border-radius:8px;padding:10px 14px;margin-bottom:4px;'>
        <div style='color:#e8b86d;font-size:0.72rem;letter-spacing:0.1em;
                    text-transform:uppercase;'>Signed in as</div>
        <div style='color:#d4e6f1;font-weight:700;font-size:0.92rem;
                    margin-top:2px;'>👤 {auth_name}</div>
        <div style='color:#5a7a96;font-size:0.74rem;'>{auth_user} · {auth_role}</div>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🚪 Sign Out", use_container_width=True, key="signout_btn"):
        for k in ["auth_ok", "auth_user", "auth_name", "auth_role", "show_user_mgmt"]:
            st.session_state.pop(k, None)
        st.rerun()
    if auth_role == "admin":
        st.markdown("---")
        if st.button("👥 Manage Users", use_container_width=True, key="btn_user_mgmt"):
            st.session_state["show_user_mgmt"] = not st.session_state.get("show_user_mgmt", False)

# Route: User Management (admin) or normal app
if auth_role == "admin" and st.session_state.get("show_user_mgmt", False):
    inject_css()
    init_db()
    seed_sample_data()
    if "delete_confirm" not in st.session_state:
        st.session_state["delete_confirm"] = False
    show_user_management(users)
else:
    main()
